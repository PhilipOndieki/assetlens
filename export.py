import io
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GOLD = "C9A84C"
NAVY = "0D1B2A"
DARK_BLUE = "1B2E45"
WHITE = "F0F4F8"
MUTED = "8899AA"
DANGER = "E74C3C"
WARNING = "F39C12"
SUCCESS = "2ECC71"
REF = "JKUAT/VAL/2025/001"


def _thin():
    s = Side(style="thin", color="243550")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, col, value, bg=DARK_BLUE, fg=GOLD, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=True, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _thin()
    return c


def _dat(ws, row, col, value, number_format=None, fg="D0D8E4", bg="0D1B2A"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", color=fg, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _thin()
    if number_format:
        c.number_format = number_format
        c.alignment = Alignment(horizontal="right", vertical="center")
    return c


def _title_row(ws, text, col_span, valuer_name, today_str):
    ws.merge_cells(f"A1:{get_column_letter(col_span)}2")
    t = ws.cell(row=1, column=1, value=text)
    t.font = Font(name="Calibri", bold=True, color=GOLD, size=16)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A3:{get_column_letter(col_span)}3")
    s = ws.cell(row=3, column=1,
                value=f"Prepared by: {valuer_name}  |  Date: {today_str}  |  Ref: {REF}")
    s.font = Font(name="Calibri", color=MUTED, size=10, italic=True)
    s.fill = PatternFill("solid", fgColor=NAVY)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18


def _sheet_register(wb, df: pd.DataFrame, valuer_name: str, today_str: str):
    ws = wb.active
    ws.title = "Asset Register"
    ws.sheet_view.showGridLines = False

    headers = ["ASSET TAG", "ASSET DESCRIPTION", "SERIAL NUMBER", "MODEL NUMBER",
               "ASSET TYPE", "LOCATION", "BUILDING", "DEPARTMENT", "FLOOR", "ROOM",
               "USER", "CONDITION", "RESERVE PRICE (KES)", "FAIR MARKET VALUE (KES)",
               "FMV/RESERVE RATIO", "IS MISSING", "IS UNTAGGED"]
    _title_row(ws, "JKUAT ASSET VALUATION REGISTER — 2025", len(headers), valuer_name, today_str)

    for ci, h in enumerate(headers, 1):
        _hdr(ws, 5, ci, h)
    ws.row_dimensions[5].height = 32

    col_map = {
        "ASSET TAG": 1, "ASSET DESCRIPTION": 2, "SERIAL NUMBER": 3, "MODEL NUMBER": 4,
        "ASSET TYPE": 5, "LOCATION": 6, "BUILDING": 7, "DEPARTMENT": 8,
        "FLOOR": 9, "ROOM": 10, "USER": 11, "CONDITION": 12,
        "RESERVE PRICE": 13, "FAIR MARKET VALUE": 14,
        "FMV_RESERVE_RATIO": 15, "IS_MISSING": 16, "IS_UNTAGGED": 17,
    }
    cond_colors = {
        "GOOD": "1A4731", "NEW": "1A4731", "SERVICEABLE": "1A4731",
        "FAIR": "4A3C00", "POOR": "4A1500",
        "FAULTY": "4A1500", "BROKEN": "3D0000",
        "OBSOLETE": "3D0000", "OLD": "3D0000",
    }

    for ri, (_, row) in enumerate(df.iterrows(), 6):
        row_bg = "111F30" if ri % 2 == 0 else "0D1B2A"
        for col_name, ci in col_map.items():
            val = row.get(col_name, "")
            if pd.isna(val):
                val = None
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", color="D0D8E4", size=10)
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _thin()
            if ci in (13, 14):
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")
            if ci == 15:
                c.number_format = "0.00"
                c.alignment = Alignment(horizontal="center", vertical="center")
            if col_name == "CONDITION":
                bg = cond_colors.get(str(val).upper(), row_bg)
                c.fill = PatternFill("solid", fgColor=bg)
            if col_name == "IS_MISSING" and val:
                c.font = Font(name="Calibri", color=DANGER, bold=True, size=10)
            if col_name == "IS_UNTAGGED" and val:
                c.font = Font(name="Calibri", color=WARNING, bold=True, size=10)

    widths = [14, 38, 18, 14, 28, 22, 42, 20, 8, 18, 16, 14, 20, 22, 14, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_summary(wb, df: pd.DataFrame, valuer_name: str, today_str: str):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    headers = ["CAMPUS", "BUILDING", "ASSET TYPE", "COUNT",
               "TOTAL FMV (KES)", "AVG FMV (KES)", "TOTAL RESERVE (KES)", "FMV/RESERVE"]
    _title_row(ws, "VALUATION SUMMARY — BY CAMPUS, BUILDING & ASSET TYPE", len(headers), valuer_name, today_str)
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 4, ci, h)
    summary = (
        df.groupby(["LOCATION", "BUILDING", "ASSET TYPE"])
        .agg(COUNT=("ASSET TAG", "count"),
             TOTAL_FMV=("FAIR MARKET VALUE", "sum"),
             AVG_FMV=("FAIR MARKET VALUE", "mean"),
             TOTAL_RESERVE=("RESERVE PRICE", "sum"))
        .reset_index()
    )
    summary["FMV_RESERVE"] = (summary["TOTAL_FMV"] / summary["TOTAL_RESERVE"]).round(2)
    for ri, (_, row) in enumerate(summary.iterrows(), 5):
        vals = [row["LOCATION"], row["BUILDING"], row["ASSET TYPE"],
                row["COUNT"], row["TOTAL_FMV"], row["AVG_FMV"],
                row["TOTAL_RESERVE"], row["FMV_RESERVE"]]
        for ci, v in enumerate(vals, 1):
            bg = "111F30" if ri % 2 == 0 else "0D1B2A"
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", color="D0D8E4", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _thin()
            if ci in (5, 6, 7):
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")
    for i, w in enumerate([22, 42, 28, 8, 20, 20, 20, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_pending(wb, pending_df: pd.DataFrame, valuer_name: str, today_str: str):
    ws = wb.create_sheet("Pending Valuation")
    ws.sheet_view.showGridLines = False
    headers = ["ASSET TAG", "ASSET DESCRIPTION", "ASSET TYPE", "LOCATION",
               "BUILDING", "DEPARTMENT", "CONDITION",
               "FAIR MARKET VALUE (KES)", "RESERVE PRICE (KES)"]
    _title_row(ws, "PENDING VALUATION — ASSETS WITHOUT FMV OR RESERVE PRICE", len(headers), valuer_name, today_str)
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 5, ci, h)
    cols = ["ASSET TAG", "ASSET DESCRIPTION", "ASSET TYPE", "LOCATION",
            "BUILDING", "DEPARTMENT", "CONDITION"]
    for ri, (_, row) in enumerate(pending_df.iterrows(), 6):
        bg = "111F30" if ri % 2 == 0 else "0D1B2A"
        for ci, col in enumerate(cols, 1):
            val = row.get(col, "") or ""
            c = ws.cell(row=ri, column=ci, value=str(val) if val else None)
            c.font = Font(name="Calibri", color="D0D8E4", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _thin()
        for ci in (8, 9):
            c = ws.cell(row=ri, column=ci, value=None)
            c.font = Font(name="Calibri", color=WARNING, size=10, italic=True)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _thin()
            c.number_format = "#,##0.00"
    for i, w in enumerate([14, 38, 28, 22, 42, 20, 14, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_missing(wb, df: pd.DataFrame, valuer_name: str, today_str: str):
    ws = wb.create_sheet("Missing Assets")
    ws.sheet_view.showGridLines = False
    missing_df = df[df["IS_MISSING"] == True]
    headers = ["ASSET TAG", "ASSET DESCRIPTION", "ASSET TYPE", "LOCATION",
               "BUILDING", "DEPARTMENT", "CONDITION",
               "FAIR MARKET VALUE (KES)", "RESERVE PRICE (KES)"]
    _title_row(ws, "MISSING ASSETS — FLAGGED DURING PHYSICAL INSPECTION", len(headers), valuer_name, today_str)
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 5, ci, h, bg="3D0000", fg=GOLD)
    cols = ["ASSET TAG", "ASSET DESCRIPTION", "ASSET TYPE", "LOCATION",
            "BUILDING", "DEPARTMENT", "CONDITION", "FAIR MARKET VALUE", "RESERVE PRICE"]
    for ri, (_, row) in enumerate(missing_df.iterrows(), 6):
        bg = "1A0000" if ri % 2 == 0 else "120000"
        for ci, col in enumerate(cols, 1):
            val = row.get(col, "") if not pd.isna(row.get(col, "")) else None
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", color="D0D8E4", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _thin()
            if ci in (8, 9):
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")
    for i, w in enumerate([14, 38, 28, 22, 42, 20, 14, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_basis(wb, valuer_name: str, today_str: str):
    ws = wb.create_sheet("Valuation Basis")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D2")
    t = ws.cell(row=1, column=1, value="VALUATION BASIS & METHODOLOGY NOTES")
    t.font = Font(name="Calibri", bold=True, color=GOLD, size=14)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    notes = [
        ("Basis of value:", "Fair Market Value — the estimated amount for which an asset should exchange on the date of valuation between a willing buyer and a willing seller."),
        ("Standards:", "Kenya Valuers Act Cap 532 and International Valuation Standards (IVS) 2022."),
        ("Data source:", "All values sourced directly from the JKUAT asset register. No estimated or synthetic values have been used."),
        ("Pending assets:", "Assets without FMV or Reserve Price are excluded from all portfolio totals and appear in the Pending Valuation sheet."),
        ("Missing assets:", "Assets flagged as not found during physical inspection appear in the Missing Assets sheet."),
        ("Condition values:", "Raw condition values are preserved as recorded: GOOD, NEW, SERVICEABLE, FAIR, POOR, FAULTY, BROKEN, OBSOLETE, OLD."),
    ]
    for ri, (label, note) in enumerate(notes, 5):
        lc = ws.cell(row=ri, column=1, value=label)
        lc.font = Font(name="Calibri", bold=True, color=GOLD, size=10)
        lc.fill = PatternFill("solid", fgColor=DARK_BLUE)
        lc.alignment = Alignment(horizontal="left", vertical="top")
        lc.border = _thin()
        ws.merge_cells(f"B{ri}:D{ri}")
        nc = ws.cell(row=ri, column=2, value=note)
        nc.font = Font(name="Calibri", color="D0D8E4", size=10)
        nc.fill = PatternFill("solid", fgColor="0D1B2A")
        nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        nc.border = _thin()
        ws.row_dimensions[ri].height = 40
    for i, w in enumerate([22, 60, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_to_excel(df: pd.DataFrame, pending_df: pd.DataFrame, valuer_name: str) -> bytes:
    wb = Workbook()
    today_str = date.today().strftime("%d %B %Y")
    _sheet_register(wb, df, valuer_name, today_str)
    _sheet_summary(wb, df, valuer_name, today_str)
    _sheet_pending(wb, pending_df, valuer_name, today_str)
    _sheet_missing(wb, df, valuer_name, today_str)
    _sheet_basis(wb, valuer_name, today_str)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_pending_excel(pending_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    today_str = date.today().strftime("%d %B %Y")
    _sheet_pending(wb, pending_df, "Philip Barongo Ondieki", today_str)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
